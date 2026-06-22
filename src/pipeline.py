from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
EXTRACT_DIR = SCRIPT_DIR.parent
RAW_DIR = EXTRACT_DIR / "raw"
DEFAULT_DATABASE = EXTRACT_DIR / "data" / "usage_stats.sqlite"
REQUIRED_SHEETS = {"API Calls": "api_calls", "Downloads": "downloads"}
CHANNELS = {"B/Ds (GNET)", "Public (Internet)"}
MONTH_NUMBERS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Incrementally load the monthly Excel workbook into SQLite.")
    parser.add_argument("--input", type=Path, help="Optional workbook path; defaults to the only .xlsx in raw/.")
    parser.add_argument("--database", type=Path, default=DEFAULT_DATABASE, help="SQLite output path.")
    parser.add_argument("--purge", action="store_true", help="Delete the selected SQLite database and exit.")
    return parser.parse_args()


def find_input_file(explicit_path: Path | None) -> Path:
    if explicit_path:
        path = explicit_path.resolve()
        if not path.is_file():
            raise ValueError(f"Input workbook not found: {path}")
        return path
    files = sorted(
        path for path in RAW_DIR.iterdir()
        if path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$")
    )
    if len(files) != 1:
        raise ValueError(f"Expected exactly one .xlsx file in {RAW_DIR}, found {len(files)}.")
    return files[0]


def purge_database(database_path: Path) -> None:
    database_path = database_path.resolve()
    if database_path.suffix.lower() != ".sqlite":
        raise ValueError(f"Refusing to purge a non-.sqlite path: {database_path}")
    targets = [
        database_path,
        Path(f"{database_path}-wal"),
        Path(f"{database_path}-shm"),
        Path(f"{database_path}-journal"),
    ]
    deleted: list[str] = []
    for target in targets:
        try:
            target.unlink()
            deleted.append(str(target))
        except FileNotFoundError:
            pass
    print(json.dumps({
        "purge": "complete",
        "database": str(database_path),
        "deleted": deleted,
        "message": "Database files deleted." if deleted else "No database files existed.",
    }, indent=2))


def cell(row: list[Any], index: int) -> Any:
    return row[index] if 0 <= index < len(row) else None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalized_text(value: Any) -> str:
    return " ".join((clean_text(value) or "").casefold().split())


def normalized_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", normalized_text(value))


def parse_month(value: Any) -> tuple[int, int] | None:
    text = clean_text(value)
    if not text:
        return None
    text = re.sub(r"\s*\*\d+\s*$", "", text)
    match = re.search(r"([A-Za-z]+)\s+(20\d{2})", text)
    if not match:
        return None
    month = MONTH_NUMBERS.get(match.group(1).lower())
    return (int(match.group(2)), month) if month else None


def find_header(rows: list[list[Any]], name: str) -> tuple[int, int] | None:
    target = normalized_header(name)
    for row_index, row in enumerate(rows[:6]):
        for column_index, value in enumerate(row):
            if normalized_header(value) == target:
                return row_index, column_index
    return None


def find_data_start(rows: list[list[Any]]) -> int:
    for row_index, row in enumerate(rows[1:], start=1):
        provider = clean_text(cell(row, 0))
        dataset = clean_text(cell(row, 1))
        data_type = normalized_text(cell(row, 2))
        if (
            provider
            and dataset
            and data_type != "total"
            and not is_annotation_provider(provider)
        ):
            return row_index
    raise ValueError("Could not locate the first dataset row.")


def is_annotation_provider(value: Any) -> bool:
    text = clean_text(value)
    if not text:
        return False
    return text.startswith("#") or re.fullmatch(r"\*\s*\d+[A-Za-z]?", text) is not None


def has_any_cell_value(row: list[Any], indexes: list[int]) -> bool:
    return any(index >= 0 and clean_text(cell(row, index)) is not None for index in indexes)


def is_dataset_row(row: list[Any], layout: dict[str, Any]) -> bool:
    provider = clean_text(cell(row, 0))
    dataset = clean_text(cell(row, 1))
    data_type = clean_text(cell(row, 2))
    if not provider or not dataset:
        return False
    if normalized_text(data_type) == "total" or is_annotation_provider(provider):
        return False
    if data_type:
        return True

    metadata_columns = [
        layout["access_group_column"],
        layout["remarks_column"],
        layout["asdp_id_column"],
    ]
    measure_columns = [dimension["column"] for dimension in layout["dimensions"]]
    return has_any_cell_value(row, metadata_columns) or has_any_cell_value(row, measure_columns)


def classify_value(raw: Any) -> tuple[int | float | None, str, str | None]:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, "missing", None
    if isinstance(raw, bool):
        return int(raw), "reported", str(raw)
    if isinstance(raw, (int, float)):
        number = int(raw) if isinstance(raw, float) and raw.is_integer() else raw
        return number, "reported", str(number)
    text = str(raw).strip()
    if text.upper() in {"N/A", "NA", "NOT APPLICABLE"}:
        return None, "not_applicable", text
    try:
        number = float(text.replace(",", ""))
        return (int(number) if number.is_integer() else number), "reported", text
    except ValueError:
        return None, "other", text


def compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def make_record_key(metric: str, identity: str, dimension: dict[str, Any]) -> str:
    source = compact_json([
        metric,
        identity,
        dimension["year"],
        dimension["month"],
        dimension["channel"] or "",
        dimension["service"] or "",
    ])
    return hashlib.sha256(source.encode("utf-8")).hexdigest()


def discover_sheet(rows: list[list[Any]], sheet_name: str, metric: str) -> dict[str, Any]:
    if (
        normalized_header(cell(rows[0], 0)) != "dataprovider"
        or normalized_header(cell(rows[0], 1)) != "datasetname"
        or normalized_header(cell(rows[0], 2)) != "datanature"
    ):
        raise ValueError(
            f"{sheet_name}: the first three headers are not Data Provider, Dataset Name, Data Nature."
        )

    data_start = find_data_start(rows)
    access_group = find_header(rows, "Access Group")
    remarks = find_header(rows, "Remarks")
    asdp_id = find_header(rows, "ASDPID")
    metadata_columns = [item[1] for item in (access_group, remarks, asdp_id) if item]
    measure_end = min(metadata_columns) if metadata_columns else max(len(row) for row in rows)
    dimensions: list[dict[str, Any]] = []
    current_date: tuple[int, int] | None = None
    current_service: str | None = None

    for column in range(3, measure_end):
        parsed_date = parse_month(cell(rows[0], column))
        if parsed_date:
            current_date = parsed_date
            current_service = None
        if not current_date:
            continue
        channel = None
        for header_row in range(1, data_start):
            header = clean_text(cell(rows[header_row], column))
            if header in CHANNELS:
                channel = header
            elif metric == "api_calls" and header:
                current_service = header
        dimensions.append({
            "column": column,
            "source_column": column + 1,
            "year": current_date[0],
            "month": current_date[1],
            "channel": channel,
            "service": current_service if metric == "api_calls" else None,
        })
    if not dimensions:
        raise ValueError(f"{sheet_name}: no month columns were discovered.")
    return {
        "data_start": data_start,
        "dimensions": dimensions,
        "access_group_column": access_group[1] if access_group else -1,
        "remarks_column": remarks[1] if remarks else -1,
        "asdp_id_column": asdp_id[1] if asdp_id else -1,
    }


def extract_records(
    rows: list[list[Any]], source_file: str, sheet_name: str, metric: str
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    layout = discover_sheet(rows, sheet_name, metric)
    records: list[dict[str, Any]] = []
    dataset_rows = 0
    for row_index, row in enumerate(rows[layout["data_start"]:], start=layout["data_start"]):
        if not is_dataset_row(row, layout):
            continue
        data_provider = clean_text(cell(row, 0))
        dataset_name = clean_text(cell(row, 1))
        data_type = clean_text(cell(row, 2))
        dataset_rows += 1
        remarks = clean_text(cell(row, layout["remarks_column"]))
        asdp_id = clean_text(cell(row, layout["asdp_id_column"]))
        access_group = clean_text(cell(row, layout["access_group_column"]))
        identity = (
            f"asdp:{normalized_text(asdp_id)}"
            if asdp_id
            else f"fallback:{normalized_text(data_provider)}|{normalized_text(dataset_name)}|{normalized_text(data_type)}"
        )
        for dimension in layout["dimensions"]:
            value, status, raw_value = classify_value(cell(row, dimension["column"]))
            records.append({
                "record_key": make_record_key(metric, identity, dimension),
                "dataset_identity": identity,
                "data_provider": data_provider,
                "dataset_name": dataset_name,
                "year": dimension["year"],
                "month": dimension["month"],
                "channel": dimension["channel"],
                "type": data_type,
                "service": dimension["service"],
                "remarks": remarks,
                "asdp_id": asdp_id,
                "metric": metric,
                "value": value,
                "value_status": status,
                "raw_value": raw_value,
                "access_group": access_group,
                "source_file": source_file,
                "source_sheet": sheet_name,
                "source_row": row_index + 1,
                "source_column": dimension["source_column"],
            })
    dates = [item["year"] * 100 + item["month"] for item in layout["dimensions"]]
    summary = {
        "sheet": sheet_name,
        "metric": metric,
        "source_rows": len(rows),
        "source_columns": max(len(row) for row in rows),
        "dataset_rows": dataset_rows,
        "dated_columns": len(layout["dimensions"]),
        "min_date": min(dates),
        "max_date": max(dates),
    }
    return records, summary


def initialize_database(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        PRAGMA journal_mode = WAL;
        PRAGMA synchronous = NORMAL;
        CREATE TABLE IF NOT EXISTS observations (
          "index" INTEGER PRIMARY KEY AUTOINCREMENT,
          data_provider TEXT NOT NULL,
          dataset_name TEXT NOT NULL,
          year INTEGER NOT NULL CHECK (year BETWEEN 2000 AND 2100),
          month INTEGER NOT NULL CHECK (month BETWEEN 1 AND 12),
          channel TEXT,
          type TEXT,
          service TEXT,
          remarks TEXT,
          asdp_id TEXT,
          metric TEXT NOT NULL CHECK (metric IN ('api_calls', 'downloads')),
          value NUMERIC,
          value_status TEXT NOT NULL CHECK (value_status IN ('reported', 'missing', 'not_applicable', 'other')),
          raw_value TEXT,
          access_group TEXT,
          source_file TEXT NOT NULL,
          source_sheet TEXT NOT NULL,
          source_row INTEGER NOT NULL,
          source_column INTEGER NOT NULL,
          dataset_identity TEXT NOT NULL,
          record_key TEXT NOT NULL UNIQUE,
          updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS import_runs (
          run_id INTEGER PRIMARY KEY AUTOINCREMENT,
          input_file TEXT NOT NULL,
          workbook_sha256 TEXT NOT NULL,
          started_at TEXT NOT NULL,
          completed_at TEXT NOT NULL,
          rows_seen INTEGER NOT NULL,
          inserted_rows INTEGER NOT NULL,
          updated_rows INTEGER NOT NULL,
          unchanged_rows INTEGER NOT NULL,
          min_date INTEGER NOT NULL,
          max_date INTEGER NOT NULL
        );
        CREATE INDEX IF NOT EXISTS observations_date_idx ON observations (year, month, "index");
        CREATE INDEX IF NOT EXISTS observations_metric_date_idx ON observations (metric, year, month);
        CREATE INDEX IF NOT EXISTS observations_asdp_idx ON observations (asdp_id);
        CREATE INDEX IF NOT EXISTS observations_provider_idx ON observations (data_provider);
        DROP VIEW IF EXISTS observations_chronological;
        CREATE VIEW observations_chronological AS
          SELECT * FROM observations ORDER BY year, month, "index";
        """
    )


SEMANTIC_FIELDS = [
    "data_provider", "dataset_name", "year", "month", "channel", "type", "service",
    "remarks", "asdp_id", "metric", "value", "value_status", "raw_value", "access_group",
    "dataset_identity",
]


def semantic_fingerprint(record: dict[str, Any] | sqlite3.Row) -> str:
    return compact_json([record[field] for field in SEMANTIC_FIELDS])


def upsert_records(
    connection: sqlite3.Connection, records: list[dict[str, Any]], run_details: dict[str, Any]
) -> dict[str, int]:
    connection.row_factory = sqlite3.Row
    existing = {
        row["record_key"]: semantic_fingerprint(row)
        for row in connection.execute(f"SELECT record_key, {', '.join(SEMANTIC_FIELDS)} FROM observations")
    }
    columns = [
        "data_provider", "dataset_name", "year", "month", "channel", "type", "service", "remarks",
        "asdp_id", "metric", "value", "value_status", "raw_value", "access_group", "source_file",
        "source_sheet", "source_row", "source_column", "dataset_identity", "record_key",
    ]
    updates = ", ".join(
        f"{column}=excluded.{column}" for column in columns if column != "record_key"
    )
    upsert_sql = f"""
        INSERT INTO observations ({', '.join(columns)})
        VALUES ({', '.join(':' + column for column in columns)})
        ON CONFLICT(record_key) DO UPDATE SET {updates}, updated_at=CURRENT_TIMESTAMP
    """
    inserted = updated = unchanged = 0
    with connection:
        for record in records:
            current = existing.get(record["record_key"])
            incoming = semantic_fingerprint(record)
            if current == incoming:
                unchanged += 1
                continue
            connection.execute(upsert_sql, record)
            if current is None:
                inserted += 1
            else:
                updated += 1
        connection.execute(
            """
            INSERT INTO import_runs (
              input_file, workbook_sha256, started_at, completed_at, rows_seen,
              inserted_rows, updated_rows, unchanged_rows, min_date, max_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_details["input_file"],
                run_details["workbook_hash"],
                run_details["started_at"],
                datetime.now(timezone.utc).isoformat(),
                len(records),
                inserted,
                updated,
                unchanged,
                run_details["min_date"],
                run_details["max_date"],
            ),
        )
    return {"inserted": inserted, "updated": updated, "unchanged": unchanged}


def main() -> None:
    options = parse_args()
    database_path = options.database.resolve()
    if options.purge:
        purge_database(database_path)
        return

    input_path = find_input_file(options.input)
    input_bytes = input_path.read_bytes()
    workbook_hash = hashlib.sha256(input_bytes).hexdigest()
    started_at = datetime.now(timezone.utc).isoformat()
    print(f"Reading {input_path.name} with openpyxl ...")
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    all_records: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []
    try:
        for sheet_name, metric in REQUIRED_SHEETS.items():
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Required worksheet not found: {sheet_name}")
            worksheet = workbook[sheet_name]
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            records, summary = extract_records(rows, input_path.name, sheet_name, metric)
            all_records.extend(records)
            summaries.append(summary)
    finally:
        workbook.close()

    all_records.sort(key=lambda record: (
        record["year"], record["month"], record["metric"], record["data_provider"],
        record["dataset_name"], record["channel"] or "", record["service"] or "",
    ))
    database_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(database_path)
    try:
        initialize_database(connection)
        min_date = min(summary["min_date"] for summary in summaries)
        max_date = max(summary["max_date"] for summary in summaries)
        changes = upsert_records(connection, all_records, {
            "input_file": input_path.name,
            "workbook_hash": workbook_hash,
            "started_at": started_at,
            "min_date": min_date,
            "max_date": max_date,
        })
        database_rows = connection.execute("SELECT COUNT(*) FROM observations").fetchone()[0]
        integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
    finally:
        connection.close()

    result = {
        "parser": f"openpyxl {sys.modules['openpyxl'].__version__}",
        "input": input_path.name,
        "database": str(database_path),
        "sheets": summaries,
        "rows_seen": len(all_records),
        **changes,
        "database_rows": database_rows,
        "date_range": [
            f"{str(min_date)[:4]}-{str(min_date)[4:]}",
            f"{str(max_date)[:4]}-{str(max_date)[4:]}",
        ],
        "integrity": integrity,
    }
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"Pipeline failed: {error}", file=sys.stderr)
        raise SystemExit(1)
